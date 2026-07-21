import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import pandas as pd
import logging
from pathlib import Path
from typing import Dict, Any
from config import REPORTS_DIR

logging.basicConfig(level=logging.INFO, format="%(asctime)s - %(levelname)s - %(message)s")

class ExcelReportExporter:
    def __init__(self, output_dir: Path = REPORTS_DIR):
        self.output_dir = output_dir
        self.output_dir.mkdir(exist_ok=True, parents=True)

        # Styling definitions
        self.header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        self.header_fill = PatternFill(start_color="1A365D", end_color="1A365D", fill_type="solid") # Dark Navy
        self.kpi_title_font = Font(name="Calibri", size=10, color="555555", bold=True)
        self.kpi_val_font = Font(name="Calibri", size=18, bold=True, color="1A365D")
        self.kpi_fill = PatternFill(start_color="F0F4F8", end_color="F0F4F8", fill_type="solid")
        self.thin_border = Border(
            left=Side(style='thin', color='D9D9D9'),
            right=Side(style='thin', color='D9D9D9'),
            top=Side(style='thin', color='D9D9D9'),
            bottom=Side(style='thin', color='D9D9D9')
        )

    def _auto_fit_columns(self, ws):
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                val_str = str(cell.value or '')
                if cell.value is not None:
                    max_len = max(max_len, len(val_str))
            ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

    def export_report(self, handle: str, 
                      summary: Dict[str, Any], 
                      topic_df: pd.DataFrame, 
                      rating_df: pd.DataFrame, 
                      diff_df: pd.DataFrame) -> Path:
        
        file_path = self.output_dir / f"Codeforces_Performance_Report_{handle}.xlsx"
        wb = openpyxl.Workbook()
        
        # -------------------------------------------------------------
        # Sheet 1: Executive Summary Dashboard
        # -------------------------------------------------------------
        ws_sum = wb.active
        ws_sum.title = "Executive Summary"
        ws_sum.views.sheetView[0].showGridLines = True

        ws_sum.cell(row=1, column=1, value=f"Codeforces Analytics Dashboard — {handle}").font = Font(name="Calibri", size=16, bold=True, color="1A365D")
        ws_sum.row_dimensions[1].height = 25

        # KPI Cards in 3x3 layout
        kpis = [
            ("Current Rating", summary.get("rating", "N/A")),
            ("Peak Rating", summary.get("max_rating", "N/A")),
            ("Rank Tier", str(summary.get("rank", "N/A")).title()),
            ("Total Contests", summary.get("total_contests", 0)),
            ("Unique Solved", summary.get("unique_solved", 0)),
            ("Total Submissions", summary.get("total_submissions", 0)),
            ("Overall Accuracy", f"{summary.get('overall_accuracy', 0)}%"),
            ("Top Strongest Topic", topic_df.iloc[0]['tag'] if not topic_df.empty else "N/A"),
            ("Weakest Accuracy Topic", topic_df.sort_values('accuracy_pct').iloc[0]['tag'] if not topic_df.empty else "N/A")
        ]

        row_start = 3
        col_start = 1
        for idx, (label, val) in enumerate(kpis):
            r = row_start + (idx // 3) * 3
            c = col_start + (idx % 3) * 2

            ws_sum.cell(row=r, column=c, value=label).font = self.kpi_title_font
            val_cell = ws_sum.cell(row=r+1, column=c, value=val)
            val_cell.font = self.kpi_val_font

            # Style block
            for r_i in range(r, r+2):
                for c_i in range(c, c+2):
                    cell = ws_sum.cell(row=r_i, column=c_i)
                    cell.fill = self.kpi_fill
                    cell.border = self.thin_border

        # -------------------------------------------------------------
        # Sheet 2: Topic Strength Matrix
        # -------------------------------------------------------------
        ws_topic = wb.create_sheet(title="Topic Matrix")
        ws_topic.views.sheetView[0].showGridLines = True
        
        headers_topic = ["Topic Tag", "Total Submissions", "Unique Attempted", "Unique Solved", "AC Submissions", "Accuracy %", "Solve Rate %", "Avg Solved Rating"]
        ws_topic.append(headers_topic)

        for col_num in range(1, len(headers_topic) + 1):
            cell = ws_topic.cell(row=1, column=col_num)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = Alignment(horizontal="center")

        if not topic_df.empty:
            for row in topic_df[['tag', 'total_submissions', 'unique_attempted', 'unique_solved', 'ac_submissions', 'accuracy_pct', 'solve_rate_pct', 'avg_solved_rating']].itertuples(index=False):
                ws_topic.append(list(row))

        self._auto_fit_columns(ws_topic)

        # -------------------------------------------------------------
        # Sheet 3: Contest History
        # -------------------------------------------------------------
        ws_rating = wb.create_sheet(title="Contest Log")
        ws_rating.views.sheetView[0].showGridLines = True

        headers_rating = ["Contest Index", "Contest ID", "Contest Name", "Date", "Rank", "Old Rating", "New Rating", "Rating Change"]
        ws_rating.append(headers_rating)

        for col_num in range(1, len(headers_rating) + 1):
            cell = ws_rating.cell(row=1, column=col_num)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = Alignment(horizontal="center")

        if not rating_df.empty:
            for row in rating_df[['contest_num', 'contest_id', 'contest_name', 'rating_update_time', 'rank', 'old_rating', 'new_rating', 'rating_change']].itertuples(index=False):
                ws_rating.append(list(row))

        self._auto_fit_columns(ws_rating)

        # -------------------------------------------------------------
        # Sheet 4: Difficulty Analysis
        # -------------------------------------------------------------
        ws_diff = wb.create_sheet(title="Difficulty Analysis")
        ws_diff.views.sheetView[0].showGridLines = True

        headers_diff = ["Rating Bin", "Total Submissions", "Unique Attempted", "Unique Solved", "AC Submissions", "Accuracy Rate %", "Solve Rate %"]
        ws_diff.append(headers_diff)

        for col_num in range(1, len(headers_diff) + 1):
            cell = ws_diff.cell(row=1, column=col_num)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = Alignment(horizontal="center")

        if not diff_df.empty:
            for row in diff_df[['rating_bin', 'total_submissions', 'unique_problems', 'unique_solved', 'ac_submissions', 'ac_rate_pct', 'solve_rate_pct']].itertuples(index=False):
                ws_diff.append(list(row))

        self._auto_fit_columns(ws_diff)

        wb.save(file_path)
        logging.info(f"Report successfully generated at: {file_path}")
        return file_path
